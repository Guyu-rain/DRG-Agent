"""DRG 规则文件解析与索引构建 (确定性, 不调用 LLM)。

参照 plans/phase1_backend.md §2.2。支持 JSON / CSV / Excel 三种格式;
解析失败不抛异常, 通过返回值的 ``parse_errors`` 字段告知调用方。
"""

from __future__ import annotations

import csv
import json
from pathlib import Path

_RULE_KEYS = ["mdc_list", "adrg_list", "drg_list", "mcc_list", "cc_list", "exclusion_table"]


def _empty_rules() -> dict:
    rules: dict = {key: [] for key in _RULE_KEYS}
    rules["parse_errors"] = []
    return rules


def _split_list(value: str | None) -> list[str]:
    """将分号/逗号分隔的字符串拆分为列表。"""
    if not value:
        return []
    return [item.strip() for item in str(value).replace("，", ",").replace(";", ",").split(",") if item.strip()]


def parse_rule_file(file_path: str | Path) -> dict:
    """解析 DRG 规则文件, 返回结构化规则字典。

    Returns:
        {
            "mdc_list": [...], "adrg_list": [...], "drg_list": [...],
            "mcc_list": [...], "cc_list": [...], "exclusion_table": [...],
            "parse_errors": [str],
        }
    """
    rules = _empty_rules()
    path = Path(file_path)

    if not path.exists():
        rules["parse_errors"].append(f"规则文件不存在: {path}")
        return rules

    suffix = path.suffix.lower()
    try:
        if suffix == ".json":
            _parse_json(path, rules)
        elif suffix == ".csv":
            _parse_csv(path, rules)
        elif suffix in (".xlsx", ".xls"):
            _parse_excel(path, rules)
        else:
            rules["parse_errors"].append(f"不支持的规则文件格式: {suffix}")
    except Exception as exc:  # noqa: BLE001 - 解析失败需优雅降级
        rules["parse_errors"].append(f"解析规则文件失败: {exc}")

    return rules


def _parse_json(path: Path, rules: dict) -> None:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        rules["parse_errors"].append("JSON 规则文件根节点必须是对象")
        return
    for key in _RULE_KEYS:
        value = data.get(key, [])
        if isinstance(value, list):
            rules[key] = value
        else:
            rules["parse_errors"].append(f"字段 {key} 应为数组")


def _parse_csv(path: Path, rules: dict) -> None:
    """CSV 为单表, 通过 rule_type 列区分规则类型。"""
    with path.open(encoding="utf-8-sig", newline="") as fh:
        for row in csv.DictReader(fh):
            rule_type = (row.get("rule_type") or "").strip().lower()
            if rule_type == "mdc":
                rules["mdc_list"].append({
                    "code": row.get("code"),
                    "name": row.get("name"),
                    "icd_prefixes": _split_list(row.get("icd_prefixes")),
                })
            elif rule_type == "adrg":
                rules["adrg_list"].append({
                    "code": row.get("code"),
                    "name": row.get("name"),
                    "mdc": row.get("mdc"),
                    "is_surgical": (row.get("is_surgical") or "").strip().lower() in ("1", "true", "yes"),
                    "surgery_list": _split_list(row.get("surgery_list")),
                    "diagnosis_list": _split_list(row.get("diagnosis_list")),
                })
            elif rule_type == "drg":
                rules["drg_list"].append({
                    "code": row.get("code"),
                    "name": row.get("name"),
                    "adrg": row.get("adrg"),
                    "cc_level": (row.get("cc_level") or "NONE").strip().upper(),
                })
            elif rule_type in ("mcc", "cc"):
                rules[f"{rule_type}_list"].append({
                    "code": row.get("code"),
                    "name": row.get("name"),
                    "level": rule_type.upper(),
                })
            elif rule_type == "exclusion":
                rules["exclusion_table"].append({
                    "diag_code": row.get("code"),
                    "excluded_by": _split_list(row.get("excluded_by")),
                })


def _parse_excel(path: Path, rules: dict) -> None:
    """Excel 每个工作表对应一种规则类型 (mdc/adrg/drg/mcc/cc/exclusion)。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        rules["parse_errors"].append("未安装 openpyxl, 无法解析 Excel 规则文件")
        return

    workbook = load_workbook(path, read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        for raw in rows[1:]:
            record = {headers[i]: raw[i] for i in range(min(len(headers), len(raw)))}
            record["rule_type"] = sheet.title.strip().lower()
            _ingest_excel_row(record, rules)
    workbook.close()


def _ingest_excel_row(record: dict, rules: dict) -> None:
    rule_type = record.get("rule_type", "")
    if rule_type == "mdc":
        rules["mdc_list"].append({
            "code": record.get("code"),
            "name": record.get("name"),
            "icd_prefixes": _split_list(record.get("icd_prefixes")),
        })
    elif rule_type == "adrg":
        rules["adrg_list"].append({
            "code": record.get("code"),
            "name": record.get("name"),
            "mdc": record.get("mdc"),
            "is_surgical": bool(record.get("is_surgical")),
            "surgery_list": _split_list(record.get("surgery_list")),
            "diagnosis_list": _split_list(record.get("diagnosis_list")),
        })
    elif rule_type == "drg":
        rules["drg_list"].append({
            "code": record.get("code"),
            "name": record.get("name"),
            "adrg": record.get("adrg"),
            "cc_level": str(record.get("cc_level") or "NONE").upper(),
        })
    elif rule_type in ("mcc", "cc"):
        rules[f"{rule_type}_list"].append({
            "code": record.get("code"),
            "name": record.get("name"),
            "level": rule_type.upper(),
        })
    elif rule_type == "exclusion":
        rules["exclusion_table"].append({
            "diag_code": record.get("code") or record.get("diag_code"),
            "excluded_by": _split_list(record.get("excluded_by")),
        })


def build_rule_index(parsed_rules: dict) -> dict:
    """将解析后的规则构建为内存索引 (哈希表), 支持 O(1) 查找。

    Returns:
        见模块文档; 关键字段含 icd_to_mdc / surgery_adrg / mcc_set / cc_set /
        exclusion_map / adrg_drg_map 等。
    """
    icd_to_mdc: dict[str, str] = {}
    mdc_names: dict[str, str] = {}
    for mdc in parsed_rules.get("mdc_list", []):
        code = mdc.get("code")
        if not code:
            continue
        mdc_names[code] = mdc.get("name", code)
        for prefix in mdc.get("icd_prefixes", []):
            icd_to_mdc[str(prefix).strip()] = code

    adrg_names: dict[str, str] = {}
    adrg_by_mdc: dict[str, list[str]] = {}
    surgery_adrg: dict[str, list[str]] = {}
    internal_adrg_by_mdc: dict[str, str] = {}
    adrg_diagnosis: dict[str, list[str]] = {}
    mdc_surgeries: dict[str, list[str]] = {}
    for adrg in parsed_rules.get("adrg_list", []):
        code = adrg.get("code")
        mdc = adrg.get("mdc")
        if not code or not mdc:
            continue
        adrg_names[code] = adrg.get("name", code)
        adrg_by_mdc.setdefault(mdc, []).append(code)
        adrg_diagnosis[code] = [str(p).strip() for p in adrg.get("diagnosis_list", [])]
        surgeries = [str(s).strip() for s in adrg.get("surgery_list", [])]
        for surgery in surgeries:
            surgery_adrg.setdefault(surgery, []).append(code)
        mdc_surgeries.setdefault(mdc, []).extend(surgeries)
        is_surgical = adrg.get("is_surgical", bool(surgeries))
        if not is_surgical and mdc not in internal_adrg_by_mdc:
            internal_adrg_by_mdc[mdc] = code

    adrg_drg_map: dict[str, list[dict]] = {}
    drg_names: dict[str, str] = {}
    for drg in parsed_rules.get("drg_list", []):
        code = drg.get("code")
        adrg = drg.get("adrg")
        if not code or not adrg:
            continue
        drg_names[code] = drg.get("name", code)
        adrg_drg_map.setdefault(adrg, []).append({
            "cc_level": str(drg.get("cc_level", "NONE")).upper(),
            "drg_code": code,
            "name": drg.get("name", code),
        })

    mcc_set = {entry.get("code") for entry in parsed_rules.get("mcc_list", []) if entry.get("code")}
    cc_set = {entry.get("code") for entry in parsed_rules.get("cc_list", []) if entry.get("code")}
    mcc_names = {e["code"]: e.get("name", e["code"]) for e in parsed_rules.get("mcc_list", []) if e.get("code")}
    cc_names = {e["code"]: e.get("name", e["code"]) for e in parsed_rules.get("cc_list", []) if e.get("code")}

    exclusion_map: dict[str, list[str]] = {}
    for entry in parsed_rules.get("exclusion_table", []):
        diag = entry.get("diag_code")
        if diag:
            exclusion_map[diag] = list(entry.get("excluded_by", []))

    return {
        "icd_to_mdc": icd_to_mdc,
        "mdc_names": mdc_names,
        "adrg_names": adrg_names,
        "adrg_by_mdc": adrg_by_mdc,
        "surgery_adrg": surgery_adrg,
        "internal_adrg_by_mdc": internal_adrg_by_mdc,
        "adrg_diagnosis": adrg_diagnosis,
        "mdc_surgeries": mdc_surgeries,
        "mcc_set": mcc_set,
        "cc_set": cc_set,
        "mcc_names": mcc_names,
        "cc_names": cc_names,
        "exclusion_map": exclusion_map,
        "adrg_drg_map": adrg_drg_map,
        "drg_names": drg_names,
    }


def count_rules(parsed_rules: dict) -> dict:
    """统计各类规则数量。"""
    return {
        "mdc": len(parsed_rules.get("mdc_list", [])),
        "adrg": len(parsed_rules.get("adrg_list", [])),
        "drg": len(parsed_rules.get("drg_list", [])),
        "mcc": len(parsed_rules.get("mcc_list", [])),
        "cc": len(parsed_rules.get("cc_list", [])),
    }
